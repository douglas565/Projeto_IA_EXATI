# back-end/excel_integration/excel_interface.py
import os
import logging
from openpyxl import load_workbook
from datetime import datetime
from ..web_capture.crawler_exati import ExatiCrawler
from ..image_processing.led_classifier import LEDClassifier
from ..cache_system.cache_manager import CacheManager
from ..data_validation.validator import LuminariaValidator

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('excel_processor.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class ExcelProcessor:
    def __init__(self):
        self.crawler = ExatiCrawler()
        self.classifier = LEDClassifier('models/led_classifier.h5')
        self.cache = CacheManager()
        self.validator = LuminariaValidator()
        
        # Configurações
        self.screenshot_dir = "screenshots"
        os.makedirs(self.screenshot_dir, exist_ok=True)

    def process_workbook(self, input_path, output_path):
        """
        Processa um arquivo Excel contendo IDs SAAC
        Colunas esperadas:
        A: ID SAAC | B: Potência (saída) | C: Tipo LED (saída) | D: Modelo (saída)
        """
        try:
            wb = load_workbook(input_path)
            sheet = wb.active
            
            # Login no EXATI (uma única vez)
            self.crawler.login()
            
            for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                id_saac = str(row[0].value).strip()
                if not id_saac:
                    continue
                
                try:
                    # Verifica cache primeiro
                    cached_data = self.cache.get(id_saac)
                    if cached_data:
                        self._fill_row(row, cached_data)
                        logger.info(f"ID {id_saac} - Dados do cache")
                        continue
                    
                    # Busca no EXATI e captura imagem
                    img_path = self.crawler.search_by_id(id_saac)
                    if not img_path:
                        raise ValueError(f"Falha ao capturar imagem para {id_saac}")
                    
                    # Processamento da imagem
                    result = self._process_image(img_path)
                    
                    # Validação cruzada
                    validated = self.validator.validate(result)
                    validated['id_saac'] = id_saac
                    validated['process_date'] = datetime.now().isoformat()
                    
                    # Preenche a planilha
                    self._fill_row(row, validated)
                    
                    # Atualiza cache
                    self.cache.set(id_saac, validated)
                    
                    logger.info(f"ID {id_saac} - Processado com sucesso")
                    
                except Exception as e:
                    logger.error(f"Erro no ID {id_saac}: {str(e)}")
                    sheet.cell(row=idx, column=5).value = f"ERRO: {str(e)}"
            
            # Salva planilha processada
            wb.save(output_path)
            logger.info(f"Planilha salva em: {output_path}")
            
        except Exception as e:
            logger.critical(f"Falha geral no processamento: {str(e)}")
            raise
        finally:
            self.crawler.close()

    def _process_image(self, image_path):
        """Executa pipeline completo de processamento de imagem"""
        try:
            # Classificação LED
            is_led = self.classifier.predict(image_path)
            
            # OCR para extração de texto
            ocr_text = self._extract_text(image_path)
            
            return {
                'image_path': image_path,
                'is_led': is_led,
                'text': ocr_text,
                'potencia': self._extract_potencia(ocr_text),
                'modelo': self._extract_modelo(ocr_text)
            }
            
        except Exception as e:
            logger.error(f"Falha no processamento da imagem {image_path}: {str(e)}")
            raise

    def _extract_text(self, image_path):
        """Extrai texto da imagem usando OCR"""
        try:
            import pytesseract
            import cv2
            
            img = cv2.imread(image_path)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            text = pytesseract.image_to_string(gray, lang='por')
            return text.strip()
            
        except Exception as e:
            logger.warning(f"OCR falhou para {image_path}: {str(e)}")
            return ""

    def _extract_potencia(self, text):
        """Extrai potência do texto usando regex"""
        import re
        match = re.search(r'(\d+)\s*W', text)
        return int(match.group(1)) if match else None

    def _extract_modelo(self, text):
        """Extrai modelo do texto"""
        import re
        match = re.search(r'Modelo[:]?\s*([A-Z0-9-]+)', text, re.IGNORECASE)
        return match.group(1) if match else "ND"

    def _fill_row(self, row, data):
        """Preenche uma linha da planilha com os dados processados"""
        row[1].value = data.get('potencia', 'N/A')      # Coluna B: Potência
        row[2].value = 'LED' if data.get('is_led') else 'Não-LED'  # Coluna C: Tipo
        row[3].value = data.get('modelo', 'ND')        # Coluna D: Modelo
        row[4].value = data.get('process_date', '')     # Coluna E: Data processamento

# Uso
if __name__ == "__main__":
    processor = ExcelProcessor()
    
    # Exemplo de execução:
    input_file = "planilha_input.xlsx"
    output_file = "planilha_processada.xlsx"
    
    processor.process_workbook(input_file, output_file)